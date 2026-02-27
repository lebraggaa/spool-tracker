import io
from datetime import datetime
from enum import Enum
from pathlib import Path

import openpyxl
import qrcode
from fastapi import FastAPI, Request, Depends, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from passlib.context import CryptContext
from sqlalchemy import (
    create_engine,
    String,
    Integer,
    DateTime,
    ForeignKey,
    Text,
    select,
    func,
)
from sqlalchemy.orm import (
    DeclarativeBase,
    mapped_column,
    Mapped,
    relationship,
    sessionmaker,
)
from starlette.middleware.sessions import SessionMiddleware

# ================= CONFIG =================

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "spool_tracker.sqlite3"
SECRET_KEY = "CHANGE_THIS_IN_PRODUCTION"

app = FastAPI(title="Spool Tracker")
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ================= DATABASE =================

engine = create_engine(
    f"sqlite:///{DB_PATH}", connect_args={"check_same_thread": False}
)
SessionLocal = sessionmaker(bind=engine)


class Base(DeclarativeBase):
    pass


# ================= ENUMS =================


class Stage(str, Enum):
    FABRICACAO = "FABRICACAO"
    LOGISTICA1 = "LOGISTICA1"
    PINTURA = "PINTURA"
    LOGISTICA2 = "LOGISTICA2"
    BORDO = "BORDO"


STAGE_ORDER = [
    Stage.FABRICACAO,
    Stage.LOGISTICA1,
    Stage.PINTURA,
    Stage.LOGISTICA2,
    Stage.BORDO,
]


def prev_stage(stage: Stage):
    i = STAGE_ORDER.index(stage)
    return STAGE_ORDER[i - 1] if i > 0 else None


class Status(str, Enum):
    PENDENTE = "PENDENTE"
    LIBERADO = "LIBERADO"
    BLOQUEADO = "BLOQUEADO"


# ================= MODELS =================


class User(Base):
    __tablename__ = "users"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    username: Mapped[str] = mapped_column(String(50), unique=True)
    password_hash: Mapped[str] = mapped_column(String(255))
    role: Mapped[str] = mapped_column(String(30))


class Spool(Base):
    __tablename__ = "spools"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    tag: Mapped[str] = mapped_column(String(120), unique=True)

    state = relationship("SpoolState", back_populates="spool", uselist=False)
    events = relationship("Event", back_populates="spool")


class SpoolState(Base):
    __tablename__ = "spool_states"

    spool_id: Mapped[int] = mapped_column(
        ForeignKey("spools.id"), primary_key=True
    )
    stage: Mapped[str] = mapped_column(String(30))
    status: Mapped[str] = mapped_column(String(20))
    location: Mapped[str | None] = mapped_column(String(100))
    note: Mapped[str | None] = mapped_column(Text)
    updated_at: Mapped[datetime] = mapped_column(DateTime)
    updated_by: Mapped[int | None] = mapped_column(ForeignKey("users.id"))

    spool = relationship("Spool", back_populates="state")


class Event(Base):
    __tablename__ = "events"

    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    spool_id: Mapped[int] = mapped_column(ForeignKey("spools.id"))
    user_id: Mapped[int | None] = mapped_column(ForeignKey("users.id"))
    ts: Mapped[datetime] = mapped_column(DateTime)
    action: Mapped[str] = mapped_column(String(30))
    stage: Mapped[str] = mapped_column(String(30))
    status: Mapped[str] = mapped_column(String(20))
    location: Mapped[str | None] = mapped_column(String(100))
    note: Mapped[str | None] = mapped_column(Text)

    spool = relationship("Spool", back_populates="events")


# ================= INIT DB =================


def init_db():
    Base.metadata.create_all(engine)
    db = SessionLocal()
    if not db.query(User).filter_by(username="admin").first():
        db.add(
            User(
                username="admin",
                password_hash=pwd_context.hash("admin123"),
                role="ADMIN",
            )
        )
        db.commit()
    db.close()


init_db()


# ================= AUTH =================


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def current_user(request: Request, db):
    uid = request.session.get("uid")
    if not uid:
        return None
    return db.query(User).get(uid)


# ================= ROUTES =================


@app.get("/", response_class=HTMLResponse)
def home(request: Request, db=Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login")

    total_blocked = (
        db.query(func.count())
        .select_from(SpoolState)
        .filter(SpoolState.status == Status.BLOQUEADO.value)
        .scalar()
    )

    return HTMLResponse(
        f"""
        <h2>Painel</h2>
        <p>Bloqueados: {total_blocked}</p>
        <a href="/spools">Buscar Spools</a>
        <br><br>
        <a href="/logout">Sair</a>
        """
    )


@app.get("/login", response_class=HTMLResponse)
def login_page():
    return HTMLResponse(
        """
        <form method="post">
        Usuário: <input name="username"><br>
        Senha: <input type="password" name="password"><br>
        <button>Entrar</button>
        </form>
        """
    )


@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...), db=Depends(get_db)):
    user = db.query(User).filter_by(username=username).first()
    if not user or not pwd_context.verify(password, user.password_hash):
        return HTMLResponse("Login inválido")

    request.session["uid"] = user.id
    return RedirectResponse("/", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login")


@app.get("/spools", response_class=HTMLResponse)
def spools(q: str = "", db=Depends(get_db)):
    query = db.query(Spool)
    if q:
        query = query.filter(Spool.tag.contains(q))

    rows = query.limit(100).all()

    html = "<h2>Spools</h2>"
    html += """
    <form>
        <input name="q">
        <button>Buscar</button>
    </form>
    """

    for s in rows:
        html += f'<a href="/spool/{s.id}">{s.tag}</a><br>'

    return HTMLResponse(html)


@app.get("/spool/{spool_id}", response_class=HTMLResponse)
def spool_detail(spool_id: int, db=Depends(get_db)):
    spool = db.query(Spool).get(spool_id)
    if not spool:
        raise HTTPException(404)

    qr_url = f"/qr/{spool.tag}.png"

    return HTMLResponse(
        f"""
        <h2>{spool.tag}</h2>
        <img src="{qr_url}" width="200"><br><br>
        <form method="post" action="/update/{spool_id}">
            Stage:
            <select name="stage">
                {''.join(f'<option>{s.value}</option>' for s in Stage)}
            </select><br>
            Status:
            <select name="status">
                {''.join(f'<option>{s.value}</option>' for s in Status)}
            </select><br>
            Localização: <input name="location"><br>
            Obs: <input name="note"><br>
            <button>Salvar</button>
        </form>
        <br><a href="/spools">Voltar</a>
        """
    )


@app.post("/update/{spool_id}")
def update_spool(spool_id: int, stage: str = Form(...), status: str = Form(...),
                 location: str = Form(""), note: str = Form(""), db=Depends(get_db)):

    spool = db.query(Spool).get(spool_id)
    if not spool:
        raise HTTPException(404)

    st = spool.state
    now = datetime.utcnow()

    if not st:
        st = SpoolState(
            spool_id=spool.id,
            stage=stage,
            status=status,
            location=location,
            note=note,
            updated_at=now,
        )
        db.add(st)
    else:
        st.stage = stage
        st.status = status
        st.location = location
        st.note = note
        st.updated_at = now

    db.add(Event(
        spool_id=spool.id,
        ts=now,
        action="UPDATE",
        stage=stage,
        status=status,
        location=location,
        note=note,
    ))

    db.commit()
    return RedirectResponse(f"/spool/{spool_id}", status_code=303)


@app.get("/qr/{tag}.png")
def qr(tag: str, request: Request):
    url = str(request.base_url) + f"spools?q={tag}"
    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return Response(content=buf.getvalue(), media_type="image/png")


@app.get("/import_excel")
def import_excel(path: str, db=Depends(get_db)):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    for r in range(1, 10):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value and "isom" in str(ws.cell(r, c).value).lower():
                col = c
                header_row = r
                break

    for row in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row, col).value
        if val:
            tag = str(val).strip()
            if not db.query(Spool).filter_by(tag=tag).first():
                db.add(Spool(tag=tag))

    db.commit()
    return {"status": "Importado com sucesso"}
